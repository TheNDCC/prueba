import io
import os
import re
import subprocess
import tkinter as tk
import unicodedata
from tkinter import messagebox

import matplotlib
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

matplotlib.use("Agg")
import matplotlib.pyplot as plt
from PIL import Image

# ==========================
#  ESTILOS REUTILIZABLES
# ==========================

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)

TOTAL_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
TOTAL_FONT = Font(bold=True, size=11)

COLUMN_WIDTHS = [18, 16, 25, 10, 18, 15, 12]

# ==========================
#  LISTA DE PRECIOS
# ==========================

PRECIOS = {
    "pollo asado entero": 340,
    "pollo asados entero": 340,
    "pollo asado medio": 170,
    "medio pollo asado": 170,
    "pollo rostizado entero": 380,
    "pollo rostizado medio": 190,
    "nachos supremos de res": 150,
    "nachos supremos mixtos": 250,
    "alitas rostizada 2 libras": 240,
    "alitas rostizada 1 libra": 120,
    "puyazo": 200,
    "churrasco": 200,
    "cerdo asado": 160,
    "carne asada": 180,
    "agua alpina 600 ml": 30,
    "gaseosa 355 ml": 30,
    "gaseosa 2 lt": 70,
    "coca cola 2lt": 70,
    "gaseosa 3 lt": 90,
    "coca cola 3lt": 90,
}


def _normalizar_texto(texto: str) -> str:
    """Elimina tildes y normaliza texto para búsqueda."""
    texto = texto.lower().strip()
    texto = unicodedata.normalize("NFD", texto)
    return "".join(c for c in texto if unicodedata.category(c) != "Mn")


def find_price(name: str):
    """Devuelve el precio según el nombre del producto usando coincidencia parcial."""
    n = _normalizar_texto(name)
    for clave, valor in PRECIOS.items():
        if clave in n or n in clave:
            return valor
    return None


# ==========================
#  FUNCIONES BASE
# ==========================


def formatear_excel(output_path: str):
    """Aplica formato a la tabla generada."""
    wb = load_workbook(output_path)
    ws = wb.active

    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    for row in range(2, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")

    for i, width in enumerate(COLUMN_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.append(["", "", "", "", "TOTAL", f"=SUM(F2:F{ws.max_row})"])
    total_row = ws.max_row
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=5)
    total_cell = ws.cell(row=total_row, column=1)
    total_cell.value = "TOTAL"
    total_cell.alignment = Alignment(horizontal="center", vertical="center")
    total_cell.fill = TOTAL_FILL
    total_cell.font = TOTAL_FONT
    total_cell.border = THIN_BORDER

    for col in range(6, ws.max_column + 1):
        cell = ws.cell(row=total_row, column=col)
        cell.fill = TOTAL_FILL
        cell.font = TOTAL_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    wb.save(output_path)


def _parse_articulos(articulos: str) -> list[tuple[int, str]]:
    """Parsea cadena de artículos (mejora: maneja comas en nombres)."""
    items = []
    for item in articulos.split(","):
        item = item.strip()
        if not item:
            continue
        match = re.match(r"^(\d+)\s+(.+)$", item)
        if match:
            cantidad = int(match.group(1))
            nombre = match.group(2).strip()
        else:
            cantidad = 1
            nombre = item.strip()
        items.append((cantidad, nombre))
    return items


def copiar_tabla_al_portapapeles(detalle: pd.DataFrame):
    """Genera una imagen de la tabla y la copia al portapapeles."""
    try:
        total = detalle["Subtotal (C$)"].sum(skipna=True)

        df_con_total = pd.concat(
            [
                detalle,
                pd.DataFrame(
                    [
                        ["TOTAL", "", "", "", "", total, ""],
                    ],
                    columns=detalle.columns,
                ),
            ],
            ignore_index=True,
        )

        fig, ax = plt.subplots(figsize=(14, max(3, len(df_con_total) * 0.5)))
        ax.axis("off")

        tabla = ax.table(
            cellText=df_con_total.values,
            colLabels=df_con_total.columns,
            cellLoc="center",
            loc="center",
            colColours=["#4472C4"] * len(df_con_total.columns),
        )

        tabla.auto_set_font_size(False)
        tabla.set_fontsize(9)
        tabla.scale(1.2, 1.5)

        num_filas = len(df_con_total)
        for key, cell in tabla.get_celld().items():
            if key[0] == 0:
                cell.set_text_props(color="white", fontweight="bold")
            elif key[0] == num_filas:
                cell.set_facecolor("#D9E2F3")
                cell.set_text_props(fontweight="bold")
            cell.set_edgecolor("#cccccc")

        plt.tight_layout()

        img_buffer = io.BytesIO()
        plt.savefig(
            img_buffer,
            format="png",
            dpi=150,
            bbox_inches="tight",
            facecolor="white",
            edgecolor="none",
        )
        plt.close(fig)

        img_buffer.seek(0)
        imagen = Image.open(img_buffer)

        temp_path = "/tmp/tabla_pedidosya.png"
        imagen.save(temp_path)

        subprocess.run(
            ["xclip", "-selection", "clipboard", "-t", "image/png", "-i", temp_path],
            check=True,
        )

        return True
    except Exception as e:
        messagebox.showwarning(
            "Advertencia",
            f"Excel generado correctamente, pero no se pudo copiar la imagen al portapapeles:\n{str(e)}",
        )
        return False


def procesar_archivo(file_path: str):
    """Lee el archivo de PedidosYa y genera un Excel con la tabla Detalle."""

    # 1) Leer CSV o Excel según extensión
    ext = os.path.splitext(file_path)[1].lower()
    try:
        if ext == ".csv":
            df = pd.read_csv(file_path, engine="python")
        elif ext in [".xls", ".xlsx"]:
            df = pd.read_excel(file_path)
        else:
            messagebox.showerror("Error", f"Tipo de archivo no soportado: {ext}")
            return
    except Exception as e:
        messagebox.showerror("Error al leer el archivo", str(e))
        return

    # 2) Verificar que existan las columnas que necesitamos
    columnas_necesarias = ["Nro de pedido", "Fecha del pedido", "Artículos"]
    for col in columnas_necesarias:
        if col not in df.columns:
            messagebox.showerror(
                "Error",
                f"En el archivo falta la columna: '{col}'.\n"
                "Verificá que el archivo sea el orderDetails original de PedidosYa.",
            )
            return

    # Estado del pedido es opcional (si no está, dejamos columna vacía)
    tiene_estado = "Estado del pedido" in df.columns

    filas = []

    for _, fila in df.iterrows():
        pedido = fila.get("Nro de pedido", "")
        fecha = fila.get("Fecha del pedido", "")
        estado = fila.get("Estado del pedido", "") if tiene_estado else ""
        articulos = fila.get("Artículos", "")

        if not isinstance(articulos, str) or not articulos.strip():
            continue

        for cantidad, nombre in _parse_articulos(articulos):
            precio = find_price(nombre)
            subtotal = cantidad * precio if precio is not None else None

            filas.append(
                {
                    "Numero de pedido": pedido,
                    "Fecha": fecha,
                    "Producto": nombre.title(),
                    "Cantidad": cantidad,
                    "Precio unitario (C$)": precio,
                    "Subtotal (C$)": subtotal,
                    "Entregado": estado,
                }
            )

    if not filas:
        messagebox.showwarning(
            "Sin datos", "No se encontraron artículos para procesar en este archivo."
        )
        return

    detalle = pd.DataFrame(filas)

    # 3) Guardar a Excel (solo hoja Detalle, como en tu imagen)
    base, ext = os.path.splitext(file_path)
    output_path = base + "_procesado.xlsx"

    try:
        with pd.ExcelWriter(output_path) as writer:
            detalle.to_excel(writer, index=False, sheet_name="Detalle")
        formatear_excel(output_path)
        copiar_tabla_al_portapapeles(detalle)
    except Exception as e:
        messagebox.showerror("Error al guardar el Excel", str(e))
        return

    # 4) Calcular total general
    total = detalle["Subtotal (C$)"].sum(skipna=True)

    messagebox.showinfo(
        "Proceso completado",
        f"Archivo generado:\n{output_path}\n\n"
        f"La tabla ha sido copiada al portapapeles como imagen.\n"
        f"Pegala en WhatsApp para compartir.\n\n"
        f"Total general: C$ {total:.2f}",
    )


# ==========================
#  INTERFAZ GRÁFICA
# ==========================


def buscar_orderdetails() -> str | None:
    """Busca orderDetails.csv en las carpetas Descargas o Downloads."""
    posibles_rutas = [
        os.path.expanduser("~/Downloads/orderDetails.csv"),
        os.path.expanduser("~/Descargas/orderDetails.csv"),
    ]

    for ruta in posibles_rutas:
        if os.path.exists(ruta):
            return ruta

    return None


def seleccionar_y_procesar():
    file_path = buscar_orderdetails()
    if file_path:
        procesar_archivo(file_path)
    else:
        messagebox.showerror(
            "Archivo no encontrado",
            "No se encontró el archivo 'orderDetails.csv' en las carpetas Descargas o Downloads.",
        )


def main():
    root = tk.Tk()
    root.title("Procesar reportes PedidosYa - Pollos Asados KM9")
    root.geometry("500x220")

    label = tk.Label(
        root,
        text=(
            "Procesador de reportes PedidosYa\n\n"
            "Se procesará automáticamente el archivo 'orderDetails.csv'\n"
            "desde las carpetas Descargas o Downloads.\n\n"
            "Se generará un Excel con la tabla Detalle:\n"
            "   Número de pedido, Fecha, Producto, Cantidad,\n"
            "   Precio unitario, Subtotal y Entregado."
        ),
        justify="left",
    )
    label.pack(pady=15)

    boton = tk.Button(
        root,
        text="Buscar y procesar orderDetails.csv",
        command=seleccionar_y_procesar,
        width=35,
        height=2,
    )
    boton.pack(pady=10)

    root.mainloop()


if __name__ == "__main__":
    main()
